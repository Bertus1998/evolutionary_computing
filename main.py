import numpy as np
import openpyxl as xl
import math
import random

from openpyxl import Workbook

ACCELERATION_MIN = 0.4
ACCELERATION_MAX = 2


def _generate_accelerate_coefs(size, phi):
    return np.random.uniform(low=0, high=phi, size=size)


def linear_interpolation(start, end, coeff):
    return (1 - coeff) * start + coeff * end


def square_interpolation(start, end, coeff):
    return end ** coeff + start ** (1 - coeff)


def const_function(start, end, coeff):
    return start + end / 2


# od 0 do 1,
def sphere_func(x: np.array):
    return np.sum(x ** 2)


# Rastrigin
def f5_func(x: np.array):
    return np.sum(x ** 2 - 10 * np.cos(math.pi * x) + 10)


def f2_func(x: np.array):
    indexes = np.arange(1, x.size + 1, 1)
    return np.sum((x - indexes) ** 2)


def griewank_func(x: np.array):
    indexes = np.arange(1, x.size + 1, 1)
    return 1 + (1 / 4000) * np.sum(x ** 2) - np.prod(np.cos(x / np.sqrt(indexes)))


def ackley_func(x: np.array):
    n = x.size
    return -20 * np.exp(-0.2 * np.sqrt((1 / n) * np.sum(x ** 2))) - np.exp(
        (1 / n) * np.sum(np.cos(2 * math.pi * x))) + 20 + math.e


def schwefel_func(x: np.array):
    return np.sum(x ** 2) + np.prod(np.abs(x))


def u(z):
    a = 10
    k = 100
    m = 4
    result = 0
    size = z.size
    for cnt in range(size):
        if z[cnt] > a:
            result = result + k * (z[cnt] - a) ** m
        elif z[cnt] < (-1) * a:
            result = result + k * ((-1) * z[cnt] * (-1) * a) ** m
        else:
            result = result + 0

    return result


def leeyao_func(x: np.array):
    n = x.size
    xi = x[0:n - 1]
    xi_plus_1 = x[1:n]
    sigma1 = np.sum(((xi - 1) ** 2) * (1 + 10 * (np.sin(math.pi * xi_plus_1)) ** 2))

    return (math.pi / n) * (10 * ((np.sin(math.pi * x[1])) ** 2) + sigma1 + (x[n - 1] - 1) ** 2) + u(x)


class ExcelSaver:
    def __init__(self, workbook: Workbook):
        # 1 - Sphere Dimension:20, 30  Range of x: ( -100, 100) Epsilon: 0.0001
        # 2 - f2 Dimension:20, 30  Range of x: ( -100, 100) Epsilon: 0.0001
        # 3 - Griewank  Dimension : 20, 30  Range of x: (-600,600) Epsilon: 0.1
        # 4 - Shwefel Dimension:20, 30  Range of x: ( -10, 10) Epsilon: 0.000001
        # 5 - LeeYao(2004) Dimension:20, 30  Range of x: ( -10, 10) Epsilon: 0.01
        self.sphere_counter_column = 1
        self.f2_counter_column = 1
        self.Griewank_counter_column = 1
        self.Shwefel_counter_column = 1
        self.LeeYao_counter_column = 1

        self.wb = workbook

    def save_result(self, worksheet, dimension, population, coefficent, result, algorithm, function):

        if function == schwefel_func:
            counter = self.Shwefel_counter_column
            self.Shwefel_counter_column += 2
        elif function == sphere_func:
            counter = self.sphere_counter_column
            self.sphere_counter_column += 2
        elif function == leeyao_func:
            counter = self.LeeYao_counter_column
            self.LeeYao_counter_column += 2
        elif function == f2_func:
            counter = self.f2_counter_column
            self.f2_counter_column += 2
        else:
            print(function)
            counter = self.Griewank_counter_column
            self.Griewank_counter_column += 2
        row_plus_1 = counter + 1
        worksheet.cell(1, counter, "dimension")  # dimension
        worksheet.cell(1, row_plus_1, "population")  # population
        worksheet.cell(2, counter, dimension)
        worksheet.cell(2, row_plus_1, population)
        worksheet.cell(3, counter, "fin_alg")  # finish algorithm
        worksheet.cell(3, row_plus_1, "acc_coe")  # coefficient acceleration
        worksheet.cell(4, counter, algorithm)
        worksheet.cell(4, row_plus_1, coefficent)
        worksheet.cell(5, counter,"value" "i")  # i
        worksheet.cell(5, row_plus_1,  "i")  # value
        worksheet.cell( 5 + result[1], counter, result[0])
        worksheet.cell( 5 + result[1], row_plus_1, result[1])
class Particle:
    def __init__(self, dimensions: int, x_from: float, x_to: float, func, ac_func):
        self.ac_func = ac_func
        self.func = func
        self.position = np.random.uniform(low=x_from, high=x_to, size=dimensions)
        self.velocity = np.random.uniform(low=x_from, high=x_to, size=dimensions)
        self.best_position = np.copy(self.position)
        self.best_score = math.inf
        self.interia_weight = 0.7

    def step(self, g, iteration_ratio):
        acc1 = ac_func(ACCELERATION_MAX, ACCELERATION_MIN, 1 - iteration_ratio)
        acc2 = ac_func(ACCELERATION_MAX, ACCELERATION_MIN, iteration_ratio)
        self.velocity = self.velocity * self.interia_weight + acc1 * random.uniform(0, 1) * (
                self.best_position - self.position) \
                        + acc2 * random.uniform(0, 1) * (g - self.position)
        self.position += self.velocity
        score = self.func(self.position)
        if score < self.best_score:
            self.best_score = score
            self.best_position = np.copy(self.position)
        return score, self.position


class Swarm:
    def __init__(self, particle_number: int, dimensions: int, x_from: float, x_to: float, func, ac_func):
        self.ac_func = ac_func
        assert particle_number > 0
        assert dimensions > 0
        self.func = func
        self.dimensions = dimensions
        self.particles = [Particle(self.dimensions, x_from, x_to, func, ac_func) for _ in range(particle_number)]
        self.best_position = np.copy(self.particles[0].best_position)
        self.best_score = math.inf

    def step(self, iteration_ratio):
        best_pos = None
        for particle in self.particles:
            score, position = particle.step(self.best_position, iteration_ratio)
            if score < self.best_score:
                self.best_score = score
                best_pos = position

        self.best_position = best_pos if best_pos is not None else self.best_position

        return self.best_score, self.best_position


class PSOAlgorithm:
    def __init__(self, swarm: Swarm, type: str, iterations: int, epsilon: float):
        self.swarm = swarm
        self.type = type
        self.iterations = iterations
        self.epsilon = epsilon

    def iteration_run(self):
        best_score = None

        for i in range(self.iterations):
            best_score, _ = self.swarm.step(i / self.iterations)
            print(f'{i}: {best_score}')
        return best_score, self.iterations

    def epsilon_run(self):

        difference = math.inf
        best_score = None
        i = 0
        for i in range(self.iterations):
            best_score, _ = self.swarm.step(i / self.iterations)
            difference = best_score
            if difference <= self.epsilon:
                break
        return best_score, self.iterations

    def run(self):
        if self.type == 'iterations':
            return self.iteration_run()
        elif self.type == 'epsilon':
            return self.epsilon_run()
        else:
            return None


if __name__ == '__main__':
    # 1 - Sphere Dimension:20, 30  Range of x: ( -100, 100) Epsilon: 0.0001
    # 2 - f2 Dimension:20, 30  Range of x: ( -100, 100) Epsilon: 0.0001
    # 3 - Griewank  Dimension : 20, 30  Range of x: (-600,600) Epsilon: 0.1
    # 4 - Shwefel Dimension:20, 30  Range of x: ( -10, 10) Epsilon: 0.000001
    # 5 - LeeYao(2004) Dimension:20, 30  Range of x: ( -10, 10) Epsilon: 0.01
    wb = Workbook()
    ws = wb.active
    excel_saver = ExcelSaver(workbook=wb)
    for i in range(5):  # funkcje celu
        if i == 0:
            epsilon = 0.0001
            low_range = -100
            high_range = 100
            function = sphere_func
            ws = wb.create_sheet("sphere_func")
        elif i == 1:
            epsilon = 0.01
            low_range = -10
            high_range = 10
            function = leeyao_func
            ws = wb.create_sheet("leeyao_func")
        elif i == 2:
            epsilon = 0.000001
            low_range = -10
            high_range = 10
            function = schwefel_func
            ws = wb.create_sheet("schwefel_func")
        elif i == 3:
            epsilon = 0.0001
            low_range = -100
            high_range = 100
            function = f2_func
            ws = wb.create_sheet("f2_func")
        else:
            epsilon = 0.1
            low_range = -600
            high_range = 600
            function = griewank_func
            ws = wb.create_sheet("griewank_func")
        for j in range(5):  # wymiar j+1 *20
            for k in range(5):  # rozmiar populacji k+1 *20
                for l in range(3):  # współczynniki przyśpieszenie
                    if l == 0:
                        ac_func = linear_interpolation
                    elif l == 1:
                        ac_func = square_interpolation
                    else:
                        ac_func = const_function
                    for m in range(2):  # dwa warianty zatrzymania algorytmu
                        for x in range(15):

                            swarm = Swarm((k + 1) * 20, (j + 1) * 20, low_range, high_range, function, ac_func)
                            if m == 0:
                                algorithm = 'iterations'
                                pso_algorithm = PSOAlgorithm(swarm, 'iterations', epsilon=epsilon, iterations=1)

                            else:
                                algorithm = 'epsilon'
                                pso_algorithm = PSOAlgorithm(swarm, 'epsilon', epsilon=epsilon, iterations=1)
                            result = pso_algorithm.run()
                            excel_saver.save_result(worksheet=ws, dimension=(j + 1) * 20, population=(k + 1) * 20,
                                                    coefficent=l,
                                                    algorithm=algorithm,
                                                    result=result, function= function)
        wb.save(filename='Test.xlsx')
# swarm = Swarm(100, 10, -10, 10, sphere_func)
# pso_algorithm = PSOAlgorithm(swarm, 'iterations', epsilon=0.01, iterations=100000)
# pso_algorithm.run()
# wb = Workbook()
#  ws = wb.active
# ws1 = wb.create_sheet("Mysheet")  # insert at the end (default)
# or
#  ws2 = wb.create_sheet("Mysheet", 0)  # insert at first position
# or
#  ws3 = wb.create_sheet("Mysheet", -1)  # insert at the penultimate position
#  ws.title = "New Title"
#  ws.sheet_properties.tabColor = "1072BA"
#  ws3 = wb["New Title"]
#  wb.save(filename='Test.xlsx')
# wcell1=ws.cell(6,1)
